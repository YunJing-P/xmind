import xmind
import openpyxl

xmind_config = {
    '用例名称': -2,
    '前置条件': [1, -3],
    '用例步骤': -3,
    '预期结果': -1
}

xlsx_config = {
    '用例名称': 1,
    '用例状态': 2,
    '用例等级': 3,
    '创建人': 4,
    '用例类型': 5,
    '用例编号': 6,
    '前置条件': 7,
    '用例步骤': 8,
    '预期结果': 9
}

default_config = {
    '用例状态': '正常',
    '用例等级': 'P2',
    '创建人': '庞允劲',
    '用例类型': '功能测试',
}


class XMind:
    def __init__(self, file_path, xmind_config, xlsx_config, default_config):
        self.file_path = file_path
        self.xmind_config = xmind_config
        self.xlsx_config = xlsx_config
        self.default_config = default_config
        self.case_min_len = 3
        for _, v in self.xmind_config.items():
            if isinstance(v, int) and abs(v) > self.case_min_len:
                self.case_min_len = abs(v)
            elif isinstance(v, list):
                if abs(v[0]) > self.case_min_len:
                    self.case_min_len = abs(v[0])
                if abs(v[1]) > self.case_min_len:
                    self.case_min_len = abs(v[1])

    def parse_xmind_and_write_xlsx(self):
        """

        :return:
        """
        xmind_book = xmind.load(self.file_path)

        n = 1
        xlsx_book = openpyxl.Workbook()
        xlsx_sheet = xlsx_book.active
        for name, col in self.xlsx_config.items():
            xlsx_sheet.cell(n, col).value = name
        n += 1

        for sheet in xmind_book.getSheets():
            root_topic = sheet.getRootTopic()
            topics_info = self.load_topics_info_flat(root_topic.getData()['topics'])
            for _, topic_info in topics_info.items():
                if topic_info['is_last']:
                    path_list = topic_info['path'][1:].split('_')

                    def merge_topic_by_list(name, config):
                        text = ''
                        num = 1
                        for topic_id in path_list[self.xmind_config[name][0]: config[name][1]]:
                            text += f'{num}.{topics_info[topic_id]["title"]}\n'
                            num += 1
                        return text

                    for config_name, default_text in self.default_config.items():
                        xlsx_sheet.cell(n, xlsx_config[config_name]).value = default_text

                    for config_name in self.xmind_config:
                        if isinstance(self.xmind_config[config_name], int):
                            xlsx_sheet.cell(n, xlsx_config[config_name]).value = \
                                topics_info[path_list[self.xmind_config[config_name]]]['title']
                        if isinstance(self.xmind_config[config_name], list):
                            xlsx_sheet.cell(n, xlsx_config[config_name]).value = merge_topic_by_list(config_name,
                                                                                                     self.xmind_config)
                    if topic_info['markers'] or topics_info[topic_info['pid']]['markers']:
                        def get_case_level(markers):
                            for marker in markers:
                                if marker.startswith('priority'):
                                    xlsx_sheet.cell(n, xlsx_config['用例等级']).value = f'P{int(marker.split("-")[1]) - 1}'
                                    break

                        if topics_info[topic_info['pid']]['markers']:
                            get_case_level(topics_info[topic_info['pid']]['markers'])
                        else:
                            get_case_level(topic_info['markers'])

                    n += 1
        xlsx_book.save('ff.xlsx')

    def load_topics_info_flat(self, topics, level=0, pid='root', path=''):
        """加载子主题列表信息，并进行扁平化处理

        :param topics: 子主题列表
        :param level: 层级，从0开始
        :param pid: 父主题id
        :param path: 全链路路径
        :return: 扁平化后的字典
        """
        topics_dict = {}
        for topic in topics:
            topics_dict[topic['id']] = {
                'id': topic['id'],
                'title': topic['title'],
                'level': level,
                'pid': pid,
                'path': f'{path}_{topic["id"]}',
                'markers': topic['markers'],
                'is_last': False,
            }
            if 'topics' in topic:
                topics_dict.update(
                    self.load_topics_info_flat(topic['topics'], level=level + 1, pid=topics_dict[topic['id']]['id'],
                                               path=topics_dict[topic['id']]['path']))
            else:
                topics_dict[topic['id']]['is_last'] = True

            if topics_dict[topic['id']]['is_last'] and topics_dict[topic['id']]['level'] < self.case_min_len - 1:
                raise IndexError(topics_dict[topic['id']]['title'])
        return topics_dict

    def write_xlsx(self):
        work_book = openpyxl.Workbook()
        work_sheet = work_book.active
        for name, col in self.xlsx_config.items():
            work_sheet.cell(1, col).value = name
        work_book.save('ff.xlsx')


p = XMind(r'C:\Users\DSAD\Desktop\xmind\实时计算平台v1.1.xmind', xmind_config, xlsx_config, default_config)
p.parse_xmind_and_write_xlsx()
